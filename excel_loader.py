import openpyxl
from openpyxl import Workbook


def getExcelFile(files, file_name):
    row_data = []
    fisrt_row = []

    wb = openpyxl.load_workbook(files[0])
    ws = wb.active
    for row in ws.iter_rows(max_row=1):
        for cell in row:
            fisrt_row.append(cell.value)
    wb.close()

    for file in files:
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            cell_data = []
            for cell in row:
                cell_data.append(cell.value)
            row_data.append(cell_data)
    wb.close()

    wr = Workbook()
    w_sheet = wr.active
    save_name = file_name + '.xlsx'
    w_sheet.title = "합친결과"

    count = 1
    for row, i in zip(fisrt_row, range(1, fisrt_row.__len__() + 1)):
        w_sheet.cell(row=count, column=i).value = row

    for row in row_data:
        count = count + 1
        num = row.__len__()
        for col, i in zip(row, range(1, num + 1)):
            save_cell = w_sheet.cell(row=count, column=i)
            save_cell.value = col
    wr.save(filename=save_name)
